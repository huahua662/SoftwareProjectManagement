import os
import logging
from typing import List, Dict, Any, AsyncGenerator
from app.models.chat_record import ChatRecord
from app.services.ai_service import AIService
from app.utils.prompt_templates import ANALYSIS_PROMPT, TASK_DECOMPOSE_PROMPT

logger = logging.getLogger(__name__)


class AnalysisService:
    def __init__(self):
        self.ai = AIService()

    _TEXT_EXTS = {".txt", ".csv", ".json", ".md", ".xml", ".html", ".htm",
                  ".yaml", ".yml", ".ini", ".cfg", ".log", ".py", ".js",
                  ".ts", ".java", ".c", ".cpp", ".h", ".sql", ".sh", ".bat"}

    def _read_file_content(self, media_path: str, max_chars: int = 8000) -> str:
        if not media_path or not os.path.isfile(media_path):
            return ""
        ext = os.path.splitext(media_path)[1].lower()

        content = ""
        try:
            if ext == ".docx":
                from docx import Document as DocxDocument
                doc = DocxDocument(media_path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                content = "\n".join(paragraphs)
            elif ext == ".pdf":
                from PyPDF2 import PdfReader
                reader = PdfReader(media_path)
                pages = []
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        pages.append(text)
                content = "\n".join(pages)
            elif ext in (".xlsx", ".xls"):
                from openpyxl import load_workbook
                wb = load_workbook(media_path, read_only=True, data_only=True)
                rows = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    if len(wb.sheetnames) > 1:
                        rows.append(f"[Sheet: {sheet}]")
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) if c is not None else "" for c in row]
                        rows.append("\t".join(cells))
                wb.close()
                content = "\n".join(rows)
            elif ext in self._TEXT_EXTS:
                with open(media_path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read(max_chars)
            else:
                return ""
        except Exception as e:
            logger.warning(f"Failed to read file {media_path}: {e}")
            return ""

        if len(content) > max_chars:
            content = content[:max_chars] + "\n...(文件内容已截断)"
        return content.strip()

    def _format_chat_context(self, records: List[ChatRecord]) -> str:
        lines = []
        for r in records:
            sender = r.sender or "未知"
            if r.msg_type == "text":
                lines.append(f"[{sender}]: {r.content}")
            elif r.msg_type in ("image", "voice", "video") and r.transcription:
                lines.append(f"[{sender}]: [{r.msg_type}] {r.transcription}")
            elif r.msg_type == "file":
                file_content = self._read_file_content(r.media_path)
                if file_content:
                    lines.append(f"[{sender}]: [文件: {r.content}]\n文件内容:\n{file_content}")
                else:
                    lines.append(f"[{sender}]: [文件: {r.content}]")
            elif r.content:
                lines.append(f"[{sender}]: {r.content}")
        return "\n".join(lines)

    async def analyze(
        self, records: List[ChatRecord], extra_prompt: str = ""
    ) -> Dict[str, Any]:
        context = self._format_chat_context(records)

        # Step 1: Generate requirement document
        analysis_messages = [
            {"role": "system", "content": ANALYSIS_PROMPT},
            {
                "role": "user",
                "content": f"以下是客户的聊天记录：\n\n{context}\n\n{extra_prompt}".strip(),
            },
        ]
        req_content = await self.ai.chat(analysis_messages, temperature=0.4, max_tokens=8192)

        # Step 2: Decompose into tasks
        task_messages = [
            {"role": "system", "content": TASK_DECOMPOSE_PROMPT},
            {
                "role": "user",
                "content": f"需求文档如下：\n\n{req_content}",
            },
        ]
        task_result = await self.ai.chat_json(task_messages, temperature=0.3, max_tokens=4096)

        tasks = task_result if isinstance(task_result, list) else task_result.get("tasks", [])

        return {
            "title": "需求分析文档",
            "content": req_content,
            "tasks": tasks,
        }

    async def analyze_stream(
        self, records: List[ChatRecord], extra_prompt: str = ""
    ) -> AsyncGenerator[Dict[str, Any], None]:
        """Stream analysis results as SSE-compatible events."""
        import json

        context = self._format_chat_context(records)

        yield {"stage": "正在分析聊天记录...", "progress": 10}

        analysis_messages = [
            {"role": "system", "content": ANALYSIS_PROMPT},
            {
                "role": "user",
                "content": f"以下是客户的聊天记录：\n\n{context}\n\n{extra_prompt}".strip(),
            },
        ]

        yield {"stage": "AI 正在生成需求文档...", "progress": 20}

        req_content = ""
        async for chunk in self.ai.chat_stream(analysis_messages, temperature=0.4, max_tokens=8192):
            req_content += chunk
            yield {"chunk": chunk}

        yield {"stage": "正在拆解开发任务...", "progress": 75}
        yield {"content": req_content}

        task_messages = [
            {"role": "system", "content": TASK_DECOMPOSE_PROMPT},
            {
                "role": "user",
                "content": f"需求文档如下：\n\n{req_content}",
            },
        ]
        task_result_text = await self.ai.chat(task_messages, temperature=0.3, max_tokens=4096)

        yield {"stage": "正在保存结果...", "progress": 90}

        task_result_text = task_result_text.strip()
        if task_result_text.startswith("```json"):
            task_result_text = task_result_text[7:]
        if task_result_text.startswith("```"):
            task_result_text = task_result_text[3:]
        if task_result_text.endswith("```"):
            task_result_text = task_result_text[:-3]
        try:
            task_result = json.loads(task_result_text.strip())
        except json.JSONDecodeError:
            task_result = {"tasks": []}

        tasks = task_result if isinstance(task_result, list) else task_result.get("tasks", [])

        yield {
            "stage": "分析完成",
            "progress": 100,
            "title": "需求分析文档",
            "final_content": req_content,
            "tasks": tasks,
            "task_count": len(tasks),
        }
